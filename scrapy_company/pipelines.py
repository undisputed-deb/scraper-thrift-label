import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

class ScrapyCompanyPipeline:
    def open_spider(self, spider):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

        # Set worksheet title and header row based on spider name
        if spider.name == "quotes":
            self.ws.title = "Quotes"
            self.headers = ["Text", "Author", "Tags"]
        elif spider.name == "thriftlabel":
            self.ws.title = "ThriftLabel"
            self.headers = ["Title", "Price", "URL"]
        else:
            self.ws.title = "Data"
            self.headers = list(spider.custom_headers) if hasattr(spider, "custom_headers") else []

        self.ws.append(self.headers)

        # Styling for headers
        bold_font = Font(bold=True)
        fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for cell in self.ws[1]:
            cell.font = bold_font
            cell.fill = fill
            cell.border = thin_border

    def process_item(self, item, spider):
        # Convert item values to match header order
        row = [", ".join(item.get(h.lower(), [])) if isinstance(item.get(h.lower()), list) else item.get(h.lower(), "") for h in self.headers]
        self.ws.append(row)
        return item

    def close_spider(self, spider):
        # Auto-adjust column widths
        for column in self.ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            self.ws.column_dimensions[column_letter].width = max_length + 2

        # Save to unique Excel file
        filename = f"{spider.name}_output.xlsx"
        self.wb.save(filename)
